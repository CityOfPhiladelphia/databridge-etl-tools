import logging 
import sys
from azure.identity import ClientSecretCredential
from msgraph_beta import GraphServiceClient
from datetime import datetime, date
import asyncio
import re
import boto3
from io import BytesIO
from openpyxl import load_workbook
import csv
import sys
import httpx

class Sharepoint():
    """Extracts a CSV from a .csv or .xlsx file in Sharepoint."""
    # if the file is .xlsx, a sheet_name must be provided to extract that sheet as csv.
    # If s3_bucket and s3_key are provided, the extracted csv will be loaded to that S3 location.
    # If not, the csv will be saved to local csv_path.
    # GraphAPI credentials must be provided to access Sharepoint.
    # The GraphAPI application must have appropriate permissions to access Sharepoint site.
    def __init__(self, 
                 graphapi_tenant_id,
                 graphapi_application_id,
                 graphapi_secret_value,
                 site_name,
                 file_path,
                 csv_path,
                 s3_bucket=None,
                 s3_key=None,
                 **kwargs):
        self.debug = kwargs.get('debug', False)

        self.graphapi_tenant_id = graphapi_tenant_id
        self.graphapi_application_id = graphapi_application_id
        self.graphapi_secret_value = graphapi_secret_value

        self.client, self.access_token = self.get_client()
        # TODO: extract the number of remaining days secret is still valid,
        # without depending on storing it in Keeper.
        # this can be done by taking in GraphAPI Object ID (NOT same as Application ID)
        # as an input argument, then calling
        # app = await self.client.applications.by_application_id(self.graphapi_object_id).get()
        # for secret in app.password_credentials:
            # if "databridge" in secret.display_name.lower():
                # self.remaining = secret.end_date_time 
                # then extract number of days from string time formatting
        self.site_name = site_name
        self.file_path = file_path
        self.file_extension = re.findall(r"(?<=\.)\w+$", self.file_path)[0]
        self.sheet_name = kwargs.get('sheet_name', None)
        self.s3_bucket = s3_bucket
        self.s3_key = s3_key
        self.csv_path = kwargs.get('csv_path', '/tmp/output.csv') # default path necessary to save before going to s3 if extracting
    
    def get_client(self) -> tuple[GraphServiceClient, str, int]:
        """Create a GraphAPI Client. Also returns the number of days for which 
        the secret is still valid (needs to be renewed every 365 days)"""
        credentials = ClientSecretCredential(
            # These field names might need to be updated if secret changes in Keeper.
            tenant_id=self.graphapi_tenant_id,
            client_id=self.graphapi_application_id,
            client_secret=self.graphapi_secret_value
        )
        token = credentials.get_token("https://graph.microsoft.com/.default")
        access_token = token.token
        if self.debug:
            print("Access token created")

        SCOPES = ['https://graph.microsoft.com/.default']
        client = GraphServiceClient(credentials=credentials, scopes=SCOPES)
        if self.debug:
            print("GraphAPI Client created")

        return client, access_token

    async def get_raw_bytes(self, url):
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Accept": "application/octet-stream"
        }
        async with httpx.AsyncClient(follow_redirects=True) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
        return response.content

    async def get_sharepoint_content(self):
        GRAPHAPI_URL_START = "https://graph.microsoft.com/v1.0/sites"
        DOMAIN = "phila.sharepoint.com"
        site_url = f"{GRAPHAPI_URL_START}/{DOMAIN}:/sites/{self.site_name}"
        site_collection_response = await self.client.sites.with_url(site_url).get()
        # actual data is contained in returned JSON as an attribute
        site = site_collection_response.additional_data
        site_id = site["id"]

        drive_url = f"{GRAPHAPI_URL_START}/{site_id}/drive"
        drive_collection_response = await self.client.sites.with_url(drive_url).get()
        drive_id = drive_collection_response.additional_data["id"]
        file_content_url = f"{drive_url}s/{drive_id}/root:/{self.file_path}:/content"
        content = await self.get_raw_bytes(file_content_url)
        return content 

    def write_to_csv(self, content):
        # TODO: make sure headers are handled properly
        if self.file_extension == "csv":
            with open(self.csv_path, "wb") as f:
                # TODO: remove all-null rows at bottom of content
                f.write(content)
        elif self.file_extension == "xlsx":
            content_stream = BytesIO(content)
            workbook = load_workbook(content_stream, read_only=True)
            if self.sheet_name is None:
                raise Exception("Must use the --sheet_name flag to specify a sheet when extracting a .xlsx file. " 
                                f"Consider one of {workbook.sheetnames}")
            else:
                try:
                    sheet = workbook[self.sheet_name]
                    if self.debug:
                        print(f"Sheet {self.sheet_name} extracted")
                except:
                    raise KeyError(f"No sheet named '{self.sheet_name} in {self.file_path}")
                with open(self.csv_path, 'w', newline='', encoding='utf-8') as f:
                    # openpyxl only provides row-by-row access to sheet data
                    for row in sheet.iter_rows(values_only=True):
                        if set([cell for cell in row]) != {None}: # remove all-null rows
                            csv.writer(f).writerow(row)
        else:
            raise Exception(f"Sharepoint file should be of type .csv or .xlsx; given {self.file_extension}")

    def load_to_s3(self):
        s3 = boto3.resource('s3')
        s3.Object(self.s3_bucket, self.s3_key).put(Body=open(self.csv_path, 'rb'))

    def extract(self):
        content = asyncio.run(self.get_sharepoint_content())
        if self.debug:
            print(f"File content of {self.file_path} obtained")
        self.write_to_csv(content)
        if self.debug:
            print(f"Content written to temporary csv at {self.csv_path}")
        
        if self.s3_bucket and self.s3_key:
            self.load_to_s3()
            if self.debug:
                print(f"Content successfully loaded to {self.s3_bucket}/{self.s3_key}")

        # TODO: consider whether to use logger instead
        # if self.remaining.days < 30: 
        #     print('GraphAPI App for Sharepoint access is expiring soon!')
        #     print(f'1. Communicate with OIT Platform Engineering to create a new certificate')
        #     print('2. Update "password" field of password manager record')
        #     print('3. Update "Expiration Date" of password manager record\n')
        #     print('Returning non-zero exit code to ensure notification; script otherwise completed successfully')
        #     sys.exit(1)

    @property
    def logger(self):
        if self._logger is None:
            logger = logging.getLogger(__name__)
            logger.setLevel(logging.INFO)
            sh = logging.StreamHandler(sys.stdout)
            logger.addHandler(sh)
            self._logger = logger
        return self._logger